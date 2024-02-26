# # import pandas as pd
# # import mysql.connector

# # # Kết nối đến cơ sở dữ liệu MySQL
# # conn = mysql.connector.connect(
# #     host='localhost',
# #     user='root',
# #     password='nghia',
# #     database='QuanLyBanHang'
# # )
# # cursor = conn.cursor()

# # # Đường dẫn đến file Excel
# # excel_file = r"/home/vvn20206205/Desktop/hht/file_input.xlsx"

# # # Đọc dữ liệu từ file Excel vào DataFrame
# # df = pd.read_excel(excel_file)

# # # Tên bảng trong cơ sở dữ liệu MySQL để lưu dữ liệu
# # table_name = 'your_table'

# # # Tạo câu lệnh SQL để chèn dữ liệu vào bảng
# # columns = ','.join(df.columns)
# # placeholders = ','.join(['%s'] * len(df.columns))
# # sql = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"

# # # Chèn dữ liệu từ DataFrame vào bảng MySQL
# # for row in df.itertuples(index=False):
# #     cursor.execute(sql, tuple(row))

# # # Commit và đóng kết nối
# # conn.commit()
# # conn.close()



# from openpyxl import load_workbook

# # Tên tệp Excel
# file_path = "your_excel_file.xlsx"

# # Tải tệp Excel vào workbook
# workbook = load_workbook(file_path)

# # Lặp qua các trang (sheets) trong workbook
# for sheet_name in workbook.sheetnames:
#     # Lấy trang hiện tại
#     sheet = workbook[sheet_name]
    
#     # In ra tên của trang
#     print("Sheet Name:", sheet_name)
    
#     # In ra nội dung của trang
#     for row in sheet.iter_rows(values_only=True):
#         print(row)
#         # tạo bảng         sheet_name.replace(" ","_") và dòng đầu tiên là tên bảng

# # Đóng workbook sau khi sử dụng
# workbook.close()
import mysql.connector
from openpyxl import load_workbook

# Kết nối đến cơ sở dữ liệu MySQL
conn = mysql.connector.connect(
    host='localhost',
    user='root',
    password='nghia',
    database='QuanLyBanHang'
)

# Tên tệp Excel
file_path = "your_excel_file.xlsx"

# Tải tệp Excel vào workbook
workbook = load_workbook(file_path)

# Lặp qua các trang (sheets) trong workbook
for sheet_name in workbook.sheetnames:
    # Lấy trang hiện tại
    sheet = workbook[sheet_name]
    
    # In ra tên của trang
    print("Sheet Name:", sheet_name)
    
    # In ra nội dung của trang
    for row in sheet.iter_rows(values_only=True):
        print(row)
        # Tạo bảng trong cơ sở dữ liệu từ nội dung của trang
        table_name = sheet_name.replace(" ", "_")
        create_table_query = f"CREATE TABLE IF NOT EXISTS {table_name} ("
        for i, column_name in enumerate(row):
            # break
            create_table_query += f"{column_name} VARCHAR(255), "
        create_table_query = create_table_query[:-2] + ")"
        
        # break
        # Thực thi truy vấn tạo bảng
        with conn.cursor() as cursor:
            cursor.execute(create_table_query)
    
# Lưu thay đổi vào cơ sở dữ liệu
conn.commit()

# Đóng workbook sau khi sử dụng
workbook.close()
